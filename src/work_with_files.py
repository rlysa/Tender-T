def save_result(file_name, *result):
    try:
        with open(file_name, encoding='utf-8', mode='w') as file:
            file.write('\n'.join(result))
        print(f'Результат сохранен в файл {file_name}')
    except Exception as e:
        print(f'Ошибка при сохранении результата в файл\n{e}')


def get_text_from_file(file_name):
    file_text = ''
    try:
        extension = file_name.split('.')[-1].lower()
        if extension == 'txt':
            with open(file_name, encoding='utf-8') as file:
                file_text = '\n'.join([i.strip() for i in file.readlines()])

        elif extension == 'xlsx':
            file_text = ''
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles.numbers import BUILTIN_FORMATS

            wb = load_workbook(file_name, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows():
                row_text = ''
                for cell in row:
                    if cell.value is not None and cell.value != 0:
                        if isinstance(cell.value, (int, float)):
                            fmt = cell.number_format
                            try:
                                row_text += str(cell.value)  if fmt == 'General' else format(cell.value, '.2f') + ' '
                            except:
                                row_text += str(cell.value) + ' '
                        else:
                            row_text += str(cell.value).lower().replace('\n', '') + ' '
                file_text += row_text + '\n'
        while '\n\n' in file_text:
            file_text = file_text.replace('\n\n', '\n')
    except Exception as e:
        print(f'Ошибка при чтении файла {file_name} \n{e}')
    return file_text


def get_text_by_words(file_text, categories):
    categories = [category.lower() for category in categories]
    category_products = {'title': ''}
    for row in file_text.split('\n'):
        if 'цена' in row or 'название' in row or 'артикул' in row or 'стоимость' in row:
            category_products['title'] += f'{row}\n'
        for category in categories:
            if category in row:
                if category not in category_products:
                    category_products[category] = ''
                category_products[category] += f'{row}\n'
    for category in category_products:
        while '\n\n' in category_products[category]:
            category_products[category] = category_products[category].replace('\n\n', '\n')
    return category_products


# if __name__ == '__main__':
#     file_text = get_text_by_words('../input/Прайс ХАТБЕР 27.08.25 цены С НДС.xlsx',  ['блокнот', 'тетрадь'])
#     title = '\n'.join(file_text['title'])
#     file_text.pop('title')
#     category_products = {}
#     products = {}
#     for category in file_text:
#         file_text[category] = '\n'.join(file_text[category])
#         from request_to_ai import make_request_to_ai
#         from prompts import prompt_get_key_info_our_products
#         answer = make_request_to_ai(prompt_get_key_info_our_products + title, file_text[category])
#         category_products[category] = []
#         for product in answer[0].strip().replace('\n\n', '\n').split('\n'):
#             if len(product.strip().split(':', 1)) == 2:
#                 article, name_cost = [i.strip() for i in product.strip().split(':', 1)]
#                 if len(name_cost.split(';', 1)) == 2:
#                     name, cost = [i.strip() for i in name_cost.split(';', 1)]
#                     category_products[category].append(article)
#                     products[article] = [name, cost]
#         print('\n\n'.join([category + '\n' + '\n'.join([product + ': ' + products[product][0] + '; ' + products[product][1] for product in category_products[category]]) for category in category_products]))
