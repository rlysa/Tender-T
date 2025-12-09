import os


def save_result(file_name, *result):
    try:
        os.makedirs(os.path.dirname(file_name), exist_ok=True)
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write('\n'.join(result))
        return True
    except Exception as e:
        raise Exception(f'Ошибка сохранения файла {file_name}: {str(e)}')


def get_text_from_file(file_name):
    file_text = ''
    try:
        extension = file_name.split('.')[-1].lower()
        if extension == 'txt':
            with open(file_name, encoding='utf-8') as file:
                file_text = '\n'.join([i.strip() for i in file.readlines()])

        elif extension == 'xlsx':
            file_text = ''
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles.numbers import BUILTIN_FORMATS

            wb = load_workbook(file_name, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows():
                row_text = ''
                for cell in row:
                    if cell.value is not None and cell != '':
                        if isinstance(cell.value, (int, float)):
                            fmt = cell.number_format
                            try:
                                row_text += str(cell.value) if fmt == 'General' else format(cell.value, '.2f') + ' '
                            except:
                                row_text += str(cell.value) + ' '
                        else:
                            row_text += str(cell.value).lower().replace('\n', '') + ' '
                if row_text.strip():
                    file_text += row_text.strip() + '\n'
        while '\n\n' in file_text:
            file_text = file_text.replace('\n\n', '\n')
        return file_text
    except Exception as e:
        raise Exception(f'Ошибка чтения файла {file_name}: {str(e)}')


def get_text_by_words(file_text, categories):
    try:
        if not file_text or not categories:
            return False
        categories = [category.lower().strip() for category in categories if category.strip()]
        category_products = {'title': ''}
        for row in file_text.split('\n'):
            if any(word in row.lower() for word in ['цена', 'название', 'артикул', 'стоимость']):
                category_products['title'] += f'{row}\n'
            for category in categories:
                if category[:-2] in row:
                    if category not in category_products:
                        category_products[category] = ''
                    category_products[category] += f'{row}\n'
                    break
        for category in category_products:
            while '\n\n' in category_products[category]:
                category_products[category] = category_products[category].replace('\n\n', '\n')
        return category_products
    except Exception as e:
        raise Exception(f"Ошибка в get_text_by_words: {str(e)}")


def update_file(path, result):
    try:
        with open(path, 'a', encoding='utf-8') as file:
            file.write(result)
        return True
    except Exception as e:
        raise Exception(f'Ошибка сохранения файла {path}: {str(e)}')

